#כדי שיעבוד יש למחוק את קובץ transactions2.xlsx

# alias - מקנה לי יכולת לקבל יכולות של ספרייה עי הקניית שם אחר לה
import openpyxl as xl

# https://pypi.org/project/openpyxl/ - דקומנטציה מפה
wb = xl.load_workbook('transactions.xlsx')
# שם הגיליון כפי שמופיע בתוך קובץ האקסל
sheet = wb['גיליון1']
# cell = sheet['a1']
cell = sheet.cell(1, 1)
# print(cell.value)

# מביא לי את כמות השורות
# print(sheet.max_row)

# for row in range(1,sheet.max_row+1):
#   print(row)

# עמ שלא יכלול את שם העמודה בגיליון האקסל נתחיל משורה שניה עד טווח מקסימלי של גיליון פלוס אחד עמ שיכלול גם את השורה האחרונה
for row in range(2, sheet.max_row + 1):
    # יתן לי את מיקום בקבצים ובאיזה גיליון
    # print(sheet.cell(row,3))

    # בסוגריים מצויין מס שורה ומס עמודה
    cell = sheet.cell(row, 3)
    # print(cell.value)
    corrected_price = cell.value * 0.9
    #מצביע על המיקום של הערכים בקובץ
    corrected_price_cell = sheet.cell(row, 4)
    #הוספת ערכים חדשים לתאים הרלוונטים שהוקצו קודם לכן
    corrected_price_cell.value = corrected_price

wb.save('transactions2.xlsx')

# You can review the names of all worksheets of the workbook with the
# Workbook.sheetname attribute this is how it goes:

# from openpyxl import load_workbook
# wb2 = load_workbook('transactions.xlsx')
# print(wb2.sheetnames)
