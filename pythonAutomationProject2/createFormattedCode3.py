import openpyxl as xl
from openpyxl.chart import BarChart, Reference


#make reusable funcation from our code till now ...
def process_workbook(fileName):
    wb = xl.load_workbook(fileName)
    sheet = wb['גיליון1']
    # cell = sheet['a1']
    cell = sheet.cell(1, 1)
    # print(cell.value)

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    # בטווח של עמודה 4 עד עמודה 4 משורה 2 עד סוף השורות בעמודה
    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4
                       )
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')
    #this is overwrite the same file
    wb.save(fileName)


